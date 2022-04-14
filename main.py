# %%
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# %%
# change quotes to fit the appropriate constellation and energystar files
energystar_excel_file = "ESPM_Excel_input\Add_Bills_to_Meters_AAPS_Historical.xlsx"
constellation_data_file = "Constellation_Excel_input\APPS_CONST_3.31.2022.xlsx"

# %%
def create_es_excel_wb(file_name):
    '''
    Inputs:
    file_name: name of energystar input file
    Returns:
    Excel workbook and excel sheet variables created with the openpyxl library
    '''
    es_workbook = load_workbook(filename=file_name)
    sheet = es_workbook["Add Bills-Non Electric"]
    return es_workbook, sheet

output_workbook, es_sheet = create_es_excel_wb(energystar_excel_file)

# %%
def delete_idxs(populated_sheet):
    '''
    Inputs:
    Energystar template excel sheet instance
    Returns:
    Energystar template excel sheet instance filtered to only include rows with our standardized naming convention.
    The convention is "Constellation__[CustomerId]__[MeterNumber]__[Meter Description]", the "__[Meter Description]" part is optional
    '''

    delete_idxs = []
    # list containing all indexes of rows that don't start with our standardized naming convention
    d_idx = 1
    for row in populated_sheet.iter_rows():
        meter_name = str(row[4].value) # extracts "Meter Name (Pre-filled)" value
        if not meter_name.startswith("Constellation__RG-"):
            # if the meter name does not start with "Constellation__RG-" (for example 009300523), then we add this index to the list
            delete_idxs.append(d_idx)
        d_idx += 1
    delete_idxs.reverse() # reverse the list so we are deleting bottom-up
    delete_idxs = delete_idxs[:-1]
    # the excel file doesn't have headers, so we skip the first row
    for idx in delete_idxs:
        # delete rows with the indexes of rows that don't match naming conventions
        populated_sheet.delete_rows(idx, amount=1)

    return populated_sheet

# %%
es_sheet = delete_idxs(es_sheet)

# %%
# this confirms that the excel file looks correct before populating it with constellation data
output_workbook.save(filename="checkpoint.xlsx")

# %%
# converts excel sheet into a dataframe, which we will later use to get the meter names
data = es_sheet.values
columns = next(data)[0:]
es_deleted_df = pd.DataFrame(data, columns=columns)

# %%
def read_energystar_data(es_df):
    '''
    Inputs:
    Energystar excel template converted into a dataframe
    Returns:
    Enegystar upload file as a dataframe, list of unique Meter names in the energystar upload template
    '''

    es_df.columns = es_df.columns.str.replace('\n', ' ')
    # gets the first part of the meter name before any commas (sometimes there would be a comma and unnessesary info after this)
    es_df["Meter Name_temp"] = es_df["Meter Name (Pre-filled)"].str.split(",").str[0]

    # this version of the regex make it so you can have any number of digits after the "RG-" in the customerID from constellation
    es_df["Meter Name_temp2"] = es_df["Meter Name_temp"][es_df["Meter Name_temp"].str.contains("Constellation__RG-\d+__\d{10}(?:.+)?") == True]
    
    es_data_meter_names = es_df["Meter Name_temp2"].unique()
    return es_df, es_data_meter_names

# %%
energystar_data, energystar_data_meter_names = read_energystar_data(es_deleted_df)

# %%
def constellation_file_cleanup(file_name):
    '''
    Inputs:
    Constellation data file name
    Returns:
    Constellation data as a dataframe, list of dataframes with each dataframe being all meter data from a particular meter
    '''
    const_excel = pd.read_excel(file_name)
    const_excel['Actual Or Estimated'] = np.where(const_excel['EndReadType'] == "Actual", 'No', 'Yes')

    # this function converts the constellation MeterNumber column into the Constellation__[CustomerId]__[MeterNumber] format
    def name_helper(row):
        row_id = str(row['CustomerId'])
        row_number = str(row['MeterNumber'])
        final_string = f"Constellation__{row_id}__{row_number}"
        return final_string
    const_excel['Const_Meter_ID'] = const_excel.apply(name_helper, axis=1)

    # this is a list of dfs, each df holds constellation data for a unique meter in constellation
    unique_meter_const = []
    for unique_meter_name in const_excel['MeterNumber'].unique():
        temp_df = const_excel[const_excel['MeterNumber'] == unique_meter_name]
        unique_meter_const.append(temp_df)

    return const_excel, unique_meter_const

# %%
constellation_excel, unique_meter_data_from_const = constellation_file_cleanup(constellation_data_file)

# %% [markdown]

# %%
# this funtion populates the energystar excel spreadsheet with the data from constellation
def populate_spreadsheet(const_dfs, es_meter_names, energystar_pop_sheet):
    '''
    Inputs:
    List of unique dataframes for constellation meters, the list of meters from the energystar template, the energystar excel sheet
    that we use for populating data
    What this does:
    Iterates through the list of constellation dataframes. Then iterates through each row in this dataframe and if constellation data
    and if the constellation meter name matches with the energystar meter name, then populate the energystar excel sheet with data from constellation
    Returns:
    Set of all unique meters and the energystar excel sheet populated with data from constellation meters
    '''
    meters_set = set()
    # this set will have all the unique meter names in the spreadsheet
    for df in const_dfs:
        # const_dfs is a list of dataframes. Each dataframe shows all of the constellation data from a unique meter
        # so looping through this list of dataframes lets us examine each unique meter from constellation at a time
        # and thus we will be adding this data to the energystar excel sheet with 1 meter from constellation's data at a time
        for row in df.iterrows():
            # row[0] is other information about the row, row[1] is the data itself
            row_data = row[1]

            if row_data['Const_Meter_ID'] not in es_meter_names:
                # if th constellation meter id is not in the list of energystar meter names, then we skip to the next dataframe
                # this means that there is a meter constellation took data on that is not in the energystar database
                continue

            # this copies the data for the row from constellation into variables like meter_number
            meter_number = row_data['MeterNumber']
            customer_id = row_data['CustomerId']
            meter_string = f"Constellation__{customer_id}__{meter_number}"

            start_date = row_data['CycleStartDate']
            end_date = row_data['CycleEndDate']
            # meter multiplier: can uncomment this later and use as needed
            # meter_multiplier = row_data['MeterMultiplier']
            fee_volume = row_data['FeeVolume']
            total_charges = row_data['TotalCharges']
            actual_estimated = row_data['Actual Or Estimated']

            # start with index 1 because excel isn't 0-based
            insert_index = 1
            for row in energystar_pop_sheet.iter_rows():
                # display(row[4].value)
                meter_name = row[4].value
                # for each row in the energystar excel sheet, we first check to see if the energystar "Meter Name (Pre-filled)" column matches the constellation meter id
                if meter_name == row_data['Const_Meter_ID']:
                    # if we the energystar meter name matches the constllation meter name, then we found the right now in the energystar excel sheet and can now inserting the constellation data into this row
                    break
                else:
                    # if the name doesn't match, then we increase the insert index number because it means we need to look at the next row in the energystar table (the meter names didn't match, so we look at the next row in constellation)
                    insert_index += 1
            
            insert_index += 1
            # incrememnt the insert index by 1 to put our new row below the row with matching meter names
            energystar_pop_sheet.insert_rows(insert_index, amount=1)
 
            # now we enter data into this new row we just created
            # there are some rows that are energystar specific (not included in the constellation reports). These are columns A, C, D, F, and J in the excel sheet
            # for these we will simply copy the information from the row above, because we know the row above is a valid row that will be accepted into energystar
            # The data from constellation that we want to transfer (like start_date, end_date, free_volumne, etc.) we enter into the matching columns in energystar 
            # ("Start Date (Required)", "End Date (Required)", "Quantity (Required)", etc.)
            energystar_pop_sheet.cell(row=insert_index, column=1).value = energystar_pop_sheet[f"A{insert_index-1}"].value
            # energystar_pop_sheet.cell(row=insert_index, column=2).value = energystar_pop_sheet[f"B{insert_index-1}"].value
            energystar_pop_sheet.cell(row=insert_index, column=3).value = energystar_pop_sheet[f"C{insert_index-1}"].value
            energystar_pop_sheet.cell(row=insert_index, column=4).value = energystar_pop_sheet[f"D{insert_index-1}"].value

            # sheet.cell(row=insert_index, column=5).value = sheet[f"E{insert_index-1}"].value
            energystar_pop_sheet.cell(row=insert_index, column=5).value = meter_string
            energystar_pop_sheet.cell(row=insert_index, column=6).value = energystar_pop_sheet[f"F{insert_index-1}"].value
            energystar_pop_sheet.cell(row=insert_index, column=7).value = start_date
            energystar_pop_sheet.cell(row=insert_index, column=8).value = end_date
            energystar_pop_sheet.cell(row=insert_index, column=9).value = fee_volume
            energystar_pop_sheet.cell(row=insert_index, column=10).value = energystar_pop_sheet[f"J{insert_index-1}"].value
            energystar_pop_sheet.cell(row=insert_index, column=11).value = total_charges
            energystar_pop_sheet.cell(row=insert_index, column=12).value = actual_estimated

            meters_set.add(meter_string)

    return (meters_set, energystar_pop_sheet)

# %%
meters_set, es_sheet = populate_spreadsheet(unique_meter_data_from_const, energystar_data_meter_names, es_sheet)

# %%
output_workbook.save(filename="output.xlsx")

# %%
# uncomment below if you want to write all of the meter names as a set and export the file:

# import json
# with open('meter_set.txt', 'w') as f:
#     f.write(str(meters_set))
